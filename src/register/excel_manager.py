import json
import openpyxl
import shutil
from utils.util_func_common import get_today_strftime_til_second
from classes.const import Category, AppMeta, FilePath, DirPath, ErrFormat
from logger import LogManager


class ExcelManager:
    logger = LogManager.get_logger()
    
    
    def __init__(self, category: str):
        # 원본 Excel, json 파일이 있는지 검사
        for file in [FilePath.EXCEL, FilePath.RESULT]:
            if not file.exists():
                raise Exception(ErrFormat.CANT_FIND.format(file.name))
        
        with open(FilePath.RESULT, "r", encoding = AppMeta.ENC_TYPE) as fp:
            self.result_dict = json.load(fp)
            self.logger.info("JSON 파일 로딩 완료")
            
        # 엑셀 파일 copy
        now = get_today_strftime_til_second()
        self.new_excel_file_path = DirPath.DATA / f"patch{now}.xlsx"
        shutil.copy(FilePath.EXCEL, self.new_excel_file_path)

        # sheet 오픈 
        sheet = category if category == Category.DOTNET.lower() else "3rdparty"
        self.excel_file = openpyxl.load_workbook(self.new_excel_file_path)
        self.excel_sheet = self.excel_file.get_sheet_by_name(sheet)
        
        # logging
        self.logger.info(f"{sheet} 시트 오픈")
        self.logger.info("엑셀 파일 로딩 완료")
    

    def get_cell_value(self, row, col: str):
        return self.excel_sheet.cell(row, self._col_to_int(col)).value


    def set_cell_value(self, row, col: str, val):
        self.excel_sheet.cell(row, self._col_to_int(col), val)


    def save_workbook(self) -> str:
        self.excel_file.save(self.new_excel_file_path)
        return self.new_excel_file_path


    def _col_to_int(self, col: str) -> int:
        val = ord(col) - ord('A') if col.isupper() else ord(col) - ord('a')
        return val + 1